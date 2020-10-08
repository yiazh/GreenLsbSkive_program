'''
Created on: 20201005

Author: Yi Zheng, Department of Electrical Engineering, DTU

'''
from equipment_package import wind_turbine, economic, converter, electrolyser
from prediction_wind_solar_price_load import ele_price_prediction
from mip import Model, xsum, MAXIMIZE, MINIMIZE
from pathlib import Path
from scipy import stats
import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import pwlf
from GLS_use_case import GreenLabSkive_Usecase1_MILP_0802
from equipment_package.math_fairy import _weibull_min, piecewise_linear
from equipment_package.math_fairy import append_df_to_excel
from openpyxl import *

# An interesting fact is if you name this file GreenLabSkive-Usecase1_MILP_0802, it can not be imported. Did you notice
# that a underscore is replaced by hyphen?

Absolute_path = Path().absolute().parent


class HWHS():
    time_span = 0.25  # h

    def __init__(self, wt=wind_turbine.wind_turbine()):
        '''

        :param wt: wind turbine object
        '''
        self.WT = wt
        print(f'The rated power of wind turbine is{wt.rated_power() * 13}')

    def linearization(self, ele_capacity=2):
        # define a converter and then conduct piecewise linearization on its power_in-power_out curve ---------
        con = converter.converter(ele_capacity)
        x = np.append(np.linspace(0, 0.1 * ele_capacity, 20), np.linspace(0.11 * ele_capacity, ele_capacity, 50))
        y = np.array([i * con.eta(i) for i in x])

        con_pwlf = pwlf.PiecewiseLinFit(x, y)
        con_breaks = con_pwlf.fit(3)  # The number indicates how many pieces are generated.
        y1 = con_pwlf.predict(x)
        # fig, ax = plt.subplots()
        # ax.plot(x, y)
        # ax.plot(x, y1)
        # plt.show()
        self.con_breaks = con_breaks
        self.con_f_breaks = np.array([i * con.eta(i) for i in con_breaks])
        self.k1_c, self.k2_c, self.k3_c = con_pwlf.slopes[0], con_pwlf.slopes[1], con_pwlf.slopes[2]
        self.b1_c, self.b2_c, self.b3_c = con_pwlf.intercepts[0], con_pwlf.intercepts[1], con_pwlf.intercepts[2]
        # ---------------------------------------End linearization of converter-------------------------------

        # do the same to electrolyser
        ele = electrolyser.electrolyser_group(I_min=0)
        y = []
        for i in x:
            electrolyser.set_power_group(ele, i)
            y.append(ele.m_H2() * 3600)
        y = np.array(y)
        # fig,ax = plt.subplots()
        # ax.plot(x,y)
        # plt.show()
        ele_pwlf = pwlf.PiecewiseLinFit(x, y)
        ele_breaks = ele_pwlf.fit(3)

        self.ele_breaks = ele_breaks

        ele_f_breaks = []
        for i in ele_breaks:
            electrolyser.set_power_group(ele, i)
            ele_f_breaks.append(ele.m_H2() * 3600)

        self.ele_f_breaks = np.array(ele_f_breaks)
        self.k1_e, self.k2_e, self.k3_e = ele_pwlf.slopes[0], ele_pwlf.slopes[1], ele_pwlf.slopes[2]
        self.b1_e, self.b2_e, self.b3_e = ele_pwlf.intercepts[0], ele_pwlf.intercepts[1], ele_pwlf.intercepts[2]
        # ---------------------------------------End linearization of electrolyser-----------------------------
        '''
        @Manual{pwlf,
            author = {Jekel, Charles F. and Venter, Gerhard},
            title = {{pwlf:} A Python Library for Fitting 1D Continuous Piecewise Linear Functions},
            year = {2019},
            url = {https://github.com/cjekel/piecewise_linear_fit_py}
                    }
        '''

    def historical_wind_power(self, start_day=0):
        '''
        Give a deterministic wind power so that we can start preliminary optimization. It would be useless in the future.
        :return: MW
        '''
        # Real data for someday in 2016
        # wind_speed = np.array([10.98, 10.81, 10.65, 10.48, 10.43, 10.37, 10.32, 10.31, 10.3, 10.29, 10.24,
        #                        10.2, 10.15, 9.89, 9.64, 9.38, 9.49, 9.61, 9.72, 9.65, 9.58, 9.5, 9.27, 9.04])
        # wind_speed = np.array([4.03, 4.02, 4.01, 4, 3.92, 3.84, 3.77, 3.91, 4.06, 4.21, 4.29, 4.36, 4.44,
        #                        4.73, 5.02, 5.31, 5.22, 5.13, 5.03, 4.94, 4.84, 4.74, 4.58, 4.42])
        File_data = Path(Absolute_path / 'prediction_wind_solar_price_load\Historical_Data\pv_wind_data_2016.csv')
        wind_speed = pd.read_csv(File_data, index_col=0)['WS10m'][start_day * 24:(start_day * 24 + 24)]

        # These are hourly data while our time span is 15 min. Thus we need to make it consistent with
        # other parameters by repeating each element four times
        return np.asarray([[13 * self.WT.wt_ac_output(i)] * 4 for i in wind_speed]).flatten()

    def historical_price(self):
        # Euros/Mwh
        return np.asarray(
            [[i] * 4 for i in ele_price_prediction.Ele_price_prediction()]).flatten()

    def historical_power_load(self):
        # MW
        return np.asarray([i + 0.8 for i in GreenLabSkive_Usecase1_MILP_0802.load_data_p_ex])

    def historical_hydrogen_load(self):
        # kg/h
        return np.asarray([i * 4 for i in GreenLabSkive_Usecase1_MILP_0802.h2_consumption_all])

    def optimize_operation_HWHS(self, number=96, tank_size=3000, ele_capacity=14, save=False,
                                linearization=False,
                                objective=0, *args, **kwargs):
        '''Operational optimization with regard to different objective function.

        :param file:
        :param linearization: True or false
        :param number: determine the time interval
        :param tank_size: kg
        :param ele_capacity: MW
        :param save: save excel or not? Bool
        :param objective: 0 for maximal profits
        :param args: reserved
        :param kwargs: a dictionary containing random variables.
        :return:
        '''
        p_w_t = kwargs['wind_power']
        pi_t = kwargs['price']
        p_l_t = kwargs['power_load']
        m_l_h2_t = kwargs['hydrogen_load']

        # marginal cost
        C_m_ele = 20.5 / 2 * 18 * 0.063 * 0.13  # Euros/MWh 1MWh electricity->20.5kg H2 ->
        # 20.5/2 kmol H2 -> 20.5/2 kmol H2O -> 20.5/2*18 kg H2O -> 0.063DKK/kg H2O
        C_m_comp = 3600 / 3.221 * 0.004

        h2_price = 2

        '''
        In fact, we need to do linearization every time when we start a new optimization, which means a new supper limit.
        However, it is very time-consuming, so I decide to do it at only one point to accelerate calculation.
        '''
        if linearization == True:
            self.ele_breaks[-1] = 20
            self.ele_f_breaks[-1] = self.k3_e * 20 + self.b3_e

            self.con_breaks[-1] =20
            self.con_f_breaks[-1] = self.k3_c * 20 + self.b3_c

        '''
        According to "Hydrogen Station Compression, Storage, and Dispensing Technical Status and Costs: Systems Integration", 
        compressor that impoves hydrogen pressure to 350bar gives an additional cost to hydrogen as $0.14/kg, provided the
        electricity cost is 1.6kWh/kg and power price is $0.085/kWh. Thus it can be infered that the addtional cost without
        electricity payment is 0.14-0.085*1.6 = 0.004$/kg

        '''
        # define MILP model
        if objective == 0 or objective == 1:
            GLS_milp_model = Model('GLS', sense=MAXIMIZE)
        else:
            GLS_milp_model = Model('GLS', sense=MINIMIZE)

        # ------------------------------------------ADD VARIABLES--------------------------------------

        power_utility = [GLS_milp_model.add_var(lb=-100, ub=100, var_type='C') for i in range(number)]
        hydrogen_level = [GLS_milp_model.add_var(lb=0, ub=tank_size, var_type='C') for i in range(number)]  # kg
        power_ele = [GLS_milp_model.add_var(lb=0, ub=ele_capacity, var_type='C') for i in
                     range(number)]
        power_comp = [GLS_milp_model.add_var(lb=0, ub=100, var_type='C') for i in range(number)]
        power_converter = [GLS_milp_model.add_var(lb=0, ub=100, var_type='C') for i in
                           range(number)]

        # variables that are only used in objective one. We need a new variable to indicate green hydrogen
        power_ele_g = [GLS_milp_model.add_var(lb=0, ub=ele_capacity, var_type='C') for _ in range(number)]

        if linearization == True:
            # Indicating which piece the electrolyser power falls in
            ele_z = []
            for i in range(3):
                ele_z.append([GLS_milp_model.add_var(var_type='B') for _ in range(number)])

            # weight
            ele_w = []
            for i in range(4):
                ele_w.append([GLS_milp_model.add_var(lb=0, var_type='C') for _ in range(number)])

            # Indicating which piece the converter input power falls in
            con_z = []
            for i in range(3):
                con_z.append([GLS_milp_model.add_var(var_type='B') for _ in range(number)])

            # weight
            con_w = []
            for i in range(4):
                con_w.append([GLS_milp_model.add_var(lb=0, var_type='C') for _ in range(number)])

        # ------------------------------------OBJECTIVE FUNCTION-----------------------------------

        Mh2_ini = 1000  # initial amount of hydrogen stored in the tank

        if objective == 0:  # Maxmize the profits
            GLS_milp_model.objective = xsum((power_utility[i] * pi_t[i] + p_l_t[i] * pi_t[i] + m_l_h2_t[i] * h2_price
                                             - power_ele[i] * C_m_ele - power_comp[i] * C_m_comp) * self.time_span
                                            for i in range(number))
        elif objective == 1:  # Maxmize percentage of green hydrogen

            GLS_milp_model.objective = xsum(power_ele_g[i] * 20.5 for i in range(number)) / sum(m_l_h2_t)

        elif objective == 2:  # Maxmize efficiency
            GLS_milp_model.objective = xsum(
                power_ele[i] * self.time_span * 3600 for i in range(number)) / (sum(m_l_h2_t) * self.time_span * 142)
            # HHV of H2 is 142MJ/kg
            # In this situation, the hydrogen production efficiency remains unchanged
            # and we can claim that a relationship between h2 production rate and electrolyser power is imperative.
            pass

        # --------------------------------------CONSTRAINS------------------------------------------
        # electrolyser power
        if linearization == True:
            for t in range(number):
                GLS_milp_model += power_ele[t] == xsum(ele_w[i][t] * self.ele_breaks[i] for i in range(4))

                GLS_milp_model += xsum(ele_z[i][t] for i in range(3)) == 1
                GLS_milp_model += xsum(ele_w[i][t] for i in range(4)) == 1

                GLS_milp_model += ele_w[0][t] <= ele_z[0][t]
                GLS_milp_model += ele_w[1][t] <= ele_z[0][t] + ele_z[1][t]
                GLS_milp_model += ele_w[2][t] <= ele_z[1][t] + ele_z[2][t]
                GLS_milp_model += ele_w[3][t] <= ele_z[2][t]

                GLS_milp_model += power_converter[t] == xsum(con_w[i][t] * self.con_breaks[i] for i in range(4))

                GLS_milp_model += xsum(con_z[i][t] for i in range(3)) == 1
                GLS_milp_model += xsum(con_w[i][t] for i in range(4)) == 1

                GLS_milp_model += con_w[0][t] <= con_z[0][t]
                GLS_milp_model += con_w[1][t] <= con_z[0][t] + con_z[1][t]
                GLS_milp_model += con_w[2][t] <= con_z[1][t] + con_z[2][t]
                GLS_milp_model += con_w[3][t] <= con_z[2][t]

                # The constrain between electrolyser power and converter power
                GLS_milp_model += power_ele[t] == xsum(con_w[i][t] * self.con_f_breaks[i] for i in range(4))

        # Green hydrogen are those produced by wind energy
        for t in range(number):
            GLS_milp_model += power_ele_g[t] <= power_ele[t]
            GLS_milp_model += power_ele_g[t] <= p_w_t[t]

        # energy conservation
        if linearization:
            for t in range(number):
                GLS_milp_model += p_w_t[t] - power_utility[t] - p_l_t[t] - power_comp[t] - power_converter[t] == 0
                pass
        else:
            for t in range(number):
                GLS_milp_model += p_w_t[t] - power_utility[t] - p_l_t[t] - power_comp[t] - power_converter[t] == 0
                pass

        # electrolyser ramping rate limit
        RLU = 10
        RLD = 10
        for t in range(number):
            if t == 0:
                GLS_milp_model += power_ele[t] - 0 <= RLU
                GLS_milp_model += power_ele[t] - 0 >= -RLD
            else:
                GLS_milp_model += power_ele[t] - power_ele[t - 1] <= RLU
                GLS_milp_model += power_ele[t] - power_ele[t - 1] >= -RLD

        # relation between compressor power and electrolyser power
        for t in range(number):
            power_comp[t] = power_ele[t] * 20.5 / 3600 * 1e3 * 3221 / 1e6  # MW, unit conversion

        # regrading hydrogen tank
        if linearization == True:
            for t in range(number):
                if t == 0:
                    GLS_milp_model += hydrogen_level[t] - Mh2_ini == (
                            xsum(self.ele_f_breaks[i] * ele_w[i][t] for i in range(4)) - m_l_h2_t[t]) * self.time_span
                else:
                    GLS_milp_model += hydrogen_level[t] - hydrogen_level[t - 1] == (
                            xsum(self.ele_f_breaks[i] * ele_w[i][t] for i in range(4)) - m_l_h2_t[t]) * self.time_span

        else:  # Assuming conversion rate of electrolyser is constant, 20.5kg/MWh
            for t in range(number):
                if t == 0:
                    GLS_milp_model += hydrogen_level[t] - Mh2_ini == (
                            power_ele[t] * 20.5 - m_l_h2_t[t]) * self.time_span
                else:
                    GLS_milp_model += hydrogen_level[t] - hydrogen_level[t - 1] == (
                            power_ele[t] * 20.5 - m_l_h2_t[t]) * self.time_span

        GLS_milp_model += hydrogen_level[-1] == Mh2_ini
        # ---------------------------------------End constrains------------------------------------

        # optimize
        GLS_milp_model.optimize()

        # ---------------------------------------SAVE RESULTS---------------------------------------
        if GLS_milp_model.status.name != 'INFEASIBLE':
            res_dict = {'Time': [i * self.time_span for i in range(number)],
                        'Price': pi_t,
                        'Power to utility grid': [round(power_utility[i].x, 2) for i in range(number)],
                        'Converter power': [round(power_converter[i].x) for i in range(number)],
                        'Electrolyser power': [round(power_ele[i].x, 2) for i in range(number)],
                        'Compressor power': [round(power_comp[i].x, 2) for i in range(number)],
                        'Wind power': p_w_t.round(2),
                        'Hydrogen level': [round(hydrogen_level[i].x) for i in range(number)],
                        'Green Electrolyser(only valid in object 1)': [round(power_ele_g[i].x, 2) for i in range(number)]
                        }
            if linearization == True:
                res_dict['mh2'] = [xsum(self.ele_f_breaks[i] * ele_w[i][t] for i in range(4)).x for t in range(number)]
            else:
                res_dict['mh2'] = [round(power_ele[t].x, 2) * 20.5 for t in range(number)]

            # -------------------------------------Daily profits---------------------------------------

            DP = np.sum([(res_dict['Power to utility grid'][i] * res_dict['Price'][i] +
                          res_dict['mh2'][i] * h2_price
                          + p_l_t[i] * res_dict['Price'][i] - res_dict['Electrolyser power'][i] * C_m_ele -
                          res_dict['Compressor power'][i] * C_m_comp) * self.time_span
                         for i in range(number)])
            DP = round(DP, 2)
            # Green hydrogen proportion
            GH2P = np.sum([min(res_dict['Electrolyser power'][i], res_dict['Wind power'][i]) for i in range(number)]) \
                   / np.sum(res_dict['Electrolyser power'])
            GH2P = round(GH2P, 2)
            #
            CR = sum(m_l_h2_t) * 142 / (np.sum(res_dict['Electrolyser power']) * 3600)
            CR = round(CR, 4)
        else:
            DP = 0
            GH2P = 0
            CR = 0

        if save == True:
            assert GLS_milp_model.status.name != 'INFEASIBLE', 'INFEASIBLE'
            save_choice = {0: 'Maximal profits_operation_',
                           1: 'Maximal green hydrogen_operation_',
                           2: 'Maximal electrolyser efficiency_operation_'}
            save_choice2 = 'Linearization' if linearization else 'No_linearization'
            file = save_choice[objective] + save_choice2

            Saving_path = Path(Path().absolute() / 'Data' / (file + '.xlsx'))
            res = pd.DataFrame(res_dict)

            try:
                res.to_excel(Saving_path, float_format='%.3f', index=False)
                print('Successfully saved')
            except PermissionError:
                print('File already exists')

            wb = load_workbook(Saving_path)
            ws = wb.active
            ws['R1'].value = 'Total amount of hydrogen/kg'
            ws['S1'].value = sum(res_dict['mh2']) * self.time_span
            ws['R2'].value = 'Total electricity consumed/MWh'
            ws['S2'].value = sum(res_dict['Electrolyser power']) * self.time_span
            ws['R3'].value = 'Conversion rate of eletrolyser/(kg/MWh)'
            ws['S3'].value = sum(res_dict['mh2']) / sum(res_dict['Electrolyser power'])
            wb.save(Saving_path)

        objective_value = GLS_milp_model.objective.x if GLS_milp_model.status.name != 'INFEASIBLE' else -1e7
        print(f'Status:{GLS_milp_model.status.name}')
        print(f'Optimal value for objective {objective}: {objective_value}')
        print(f'Daily profits: {DP}')
        print(f'Green hydrongen proportion: {GH2P}')
        print(f'Average efficiency of electrolyser:{CR}')

        self.status = GLS_milp_model.status.name
        self.objective_value = objective_value
        self.DP = DP
        self.GH2P = GH2P
        self.CR = CR

        return GLS_milp_model.status.name, objective_value, DP

    def sensitivity_HWHS(self, nodes=20, objective=0, max_ele=10, max_tank=2000, figure=False, linearization=False):
        # A sensitivity analysis on design parameters: size of tank and electrolyser capacity

        capacity_ele = [i for i in np.linspace(2, max_ele, nodes)]
        tank_size = [i for i in np.linspace(1000, max_tank, nodes)]
        inner_result = np.zeros((nodes, nodes))

        min_ele = 1e3  # The minimal electrolyser capacity that can make the problem solvable.
        min_tank = 1e7  #
        if figure == True:
            fig = plt.figure()
            ax = fig.add_subplot(111, projection='3d')
            ax.set_xlabel('Capacity of electrolyser(MW)')
            ax.set_ylabel('Tank size(kg)')
            if objective == 0:
                ax.set_zlabel('Optimal operational profit(€)')
            elif objective == 1:
                ax.set_zlabel('Optimzal green hydrogen proportion')
            elif objective == 2:
                ax.set_zlabel('Average hydrogen conversion rate')
        for i, ele in enumerate(capacity_ele):
            for j, tank in enumerate(tank_size):
                status, optimized_result, dp = self.optimize_operation_HWHS(tank_size=tank, ele_capacity=ele,
                                                                            objective=objective,
                                                                            wind_power=self.historical_wind_power(),
                                                                            price=self.historical_price(),
                                                                            power_load=self.historical_power_load(),
                                                                            hydrogen_load=self.historical_hydrogen_load(),
                                                                            save=False,
                                                                            linearization=linearization
                                                                            )
                if status != 'INFEASIBLE':
                    if objective != 2:
                        inner_result[(i, j)] = optimized_result
                    else:
                        inner_result[(i, j)] = self.CR
                    if figure:
                        ax.scatter(ele, tank, inner_result[(i, j)], marker='.', color='grey')
                    min_ele = ele if min_ele > ele else min_ele
                    min_tank = tank if min_tank > tank else min_tank
                else:
                    inner_result[(i, j)] = 0
                print(f'Cycle: ele_capacity ={ele}, tank_size = {tank}')
        if figure == True:
            plt.show()
        return min_ele, min_tank

    def IRR(self, ele_capacity=14, tank_size=3000):
        '''Given a specific electrolyser capacity and tank size, this function returns the system's Internal rate of return.

        :param ele_capacity: MW
        :param tank_size: kg
        :return: IRR \in (-1,1)
        '''
        interest_rate = 0.07  # for estimating the replacement cost. It's better to precisely calculate the cost but as this
        # cost is relatively small, we can simply ignore it.

        c_cap = 13 * self.WT.rated_power() * self.WT.capital_cost * 1000 + 1000 * ele_capacity * 1492 \
                + tank_size * 854 + 1000 * ele_capacity * 126 + ele_capacity * 20.5 * 13338
        c_om = 13 * self.WT.rated_power() * 1000 * 56 + 1000 * ele_capacity * 60 + tank_size * 8 + 20.5 * ele_capacity * 666
        c_rep = 1000 * ele_capacity * 126 / (math.pow(1 + interest_rate, 15))

        # Only after operational dispatch can we decide the daily profits
        p_annual = self.DP

        print(f'The annual cost and annual profit are {c_om} , {365 * p_annual}, respectively ')

        irr = economic.internal_rate_of_return(initial_investment=c_cap + c_rep, annual_cost=c_om,
                                               annual_profit=p_annual * 365, length_project=20)
        print(f'IRR is {irr}')
        return irr

        # # Invoke particle swarm optimization algorithm to solve this maxmial problem.
        # def objective_function(x):
        #     return -IRR(ele_capacity=x[0], tank_size=x[1])
        #
        # lb = [10, 1000]
        # ub = [20, 6000]
        # xopt,fopt = pso(objective_function, lb, ub,swarmsize=100, maxiter=200, omega=0.5)
        # return xopt,fopt


class Uncertainity_variable():
    # wind speed
    Saving_path = Absolute_path / 'Hybrid_wind_hydrogen/Data' / 'UV.xlsx'

    def wind_speed(self):
        # The meterologicl data is saved in this file. Caution: when I am using Path().absolute() to get the path, the
        # return value differs according to where I provoke this funtion
        data_path = Absolute_path / 'prediction_wind_solar_price_load/Historical_Data/pv_wind_data_2016.csv'
        WS_10_m = pd.read_csv(data_path)['WS10m']

        # Fitting the speed data with weibull distribution
        # For some reason, the argument 'loc' has to be set to a value. If not, the fitting would be very poor.
        # For details, check https://github.com/scipy/scipy/issues/11806
        shape, loc, scale = stats.weibull_min.fit(WS_10_m, loc=7)
        print(shape, loc, scale)

        # -------------------------------------plot statistical results----------------------------
        # Interval number
        num_bins = 100

        # Dipict frequency distribution histgram
        fig, ax = plt.subplots()
        # N is the count in each bin, bins is the lower-limit of the bin
        N, bins, patches = ax.hist(WS_10_m, num_bins)

        # ------------------------------------Fitting curve-----------------------------------------
        fig1, ax1 = plt.subplots()
        x = np.linspace(WS_10_m.min(), WS_10_m.max(), num_bins)
        y1 = [stats.weibull_min.pdf(i, shape, loc, scale) for i in x]
        ax1.plot(x, y1)
        # Self-defined function
        y2 = [_weibull_min(i, shape, loc, scale) for i in x]
        ax1.plot(x, y2, color='red')

        # ------------------------------------Fitting curve and statistical data--------------------
        fig2, ax2 = plt.subplots()
        ax2.hist(WS_10_m, num_bins, density=True, color='tab:blue', alpha=0.5)
        ax2.plot(x, y1, label='Fitted weibull dist')
        ax2.set(xlabel='Wind speed(m/s)', ylabel='Frequency', title='Weibull distribution fitting')
        ax2.legend()
        plt.show()

        # Save data
        wind_speed_data = pd.DataFrame()
        bins = bins.tolist()
        bins.pop(-1)
        wind_speed_data['Speed'] = bins
        wind_speed_data['Frequency'] = [i / WS_10_m.size for i in N]
        wind_speed_data['Speed_2'] = x
        wind_speed_data['Weibull_pdf'] = y1
        try:
            wind_speed_data.to_excel(Uncertainity_variable.Saving_path, sheet_name='ws', float_format='%.3f',
                                     index=False)
        except PermissionError:
            print('File is open')

    def electricity_price(self):
        data_path = Absolute_path / 'prediction_wind_solar_price_load/Historical_Data/elspot-prices_2020_hourly_eur1.csv'
        ele_price = pd.read_csv(data_path)['DK2']
        # -------------------------------------print statistical results----------------------------
        # Interval number
        num_bins = 150

        # Dipict frequency distribution histgram
        fig, ax = plt.subplots()
        # N is the count in each bin, bins is the lower-limit of the bin
        N, bins, patches = ax.hist(ele_price, num_bins)
        a = np.asarray(ele_price)
        for index, value in enumerate(a):
            if np.isfinite(value) == False:
                print(index)

        # -------------------------------------Fitting -----------------------------------------------
        loc, scale = stats.norm.fit(ele_price, loc=10)
        print(f'Parameters of normal distribution is mu = {round(loc, 2)},sigma = {round(scale, 2)}')
        fig1, ax1 = plt.subplots()
        x = np.linspace(ele_price.min(), ele_price.max(), num_bins)
        '''
        Specifically, norm.pdf(x, loc, scale) is identically equivalent to 
        norm.pdf(y) / scale with y = (x - loc) / scale
        '''
        y1 = [stats.norm.pdf(i, loc, scale) for i in x]
        ax1.plot(x, y1, label='Fitted normal dist')
        ax1.hist(ele_price, num_bins, color='tab:orange', alpha=0.5, density=True)
        ax1.set(xlabel='Price/€', ylabel='Frequency')
        ax1.legend()
        fig1.tight_layout()
        plt.show()

    def hydrogen_demand(self):
        pass


# Invoke jmetal package to solve this multi-objective optmizaiton problem.
from jmetal.core.problem import FloatProblem
from jmetal.core.solution import FloatSolution


class sizing_hwhs_problem(FloatProblem):
    def __init__(self, hwhs: HWHS, linearization=False, objective=0):
        super(sizing_hwhs_problem, self).__init__()
        self.number_of_variables = 2
        self.number_of_objectives = 2
        self.number_of_constraints = 0

        self.obj_directions = [self.MAXIMIZE, self.MAXIMIZE]
        self.obj_labels = ['x', 'y']

        min_ele, min_tank = hwhs.sensitivity_HWHS(objective=objective)

        self.linearization = linearization
        self.objective = objective
        if self.linearization:
            hwhs.linearization(ele_capacity=14)

        self.lower_bound = [min_ele + 1, min_tank + 500]
        self.upper_bound = [20, 6000]

        FloatSolution.lower_bound = self.lower_bound
        FloatSolution.upper_bound = self.upper_bound

        self.hwhs = hwhs

    def evaluate(self, solution: FloatSolution) -> FloatSolution:
        ele_capacity = solution.variables[0]
        tank_size = solution.variables[1]
        self.hwhs.optimize_operation_HWHS(tank_size=tank_size,
                                          ele_capacity=ele_capacity,
                                          objective=self.objective,
                                          wind_power=self.hwhs.historical_wind_power(),
                                          price=self.hwhs.historical_price(),
                                          power_load=self.hwhs.historical_power_load(),
                                          hydrogen_load=self.hwhs.historical_hydrogen_load(),
                                          save=False,
                                          linearization=self.linearization
                                          )
        if self.objective == 2:
            solution.objectives[0] = self.hwhs.CR
        elif self.objective == 1:
            solution.objectives[0] = self.hwhs.GH2P
        else:
            solution.objectives[0] = self.hwhs.DP
        solution.objectives[1] = self.hwhs.IRR(ele_capacity=ele_capacity, tank_size=tank_size)

        return solution

    def __evaluate_constraints(self, solution: FloatSolution) -> None:  # remained
        pass

    def get_name(self) -> str:
        return 'sizing_hwhs_problem'


if __name__ == '__main__':
    a = HWHS(wind_turbine.wind_turbine(r=45, height=55))
    a.linearization(ele_capacity=15)
    a.optimize_operation_HWHS(objective=2,
                              wind_power=a.historical_wind_power(),
                              price=a.historical_price(),
                              power_load=a.historical_power_load(),
                              hydrogen_load=a.historical_hydrogen_load(),
                              save=True,
                              linearization=True,
                              ele_capacity=14
                              )
    a.sensitivity_HWHS(objective=2, figure=True, max_ele=20, max_tank=4000, linearization=True)
    # c, d = a.sensitivity_HWHS()
    # print(f'min_ele:{c}')
    # print(d)
